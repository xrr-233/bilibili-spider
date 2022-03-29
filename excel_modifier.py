import openpyxl

if(__name__=="__main__"):
    wb = openpyxl.load_workbook('修正数据.xlsx')
    ws = wb['Sheet1']
    for i in range(2, ws.max_row + 1):
        if(str(ws.cell(row=i, column=7).value)[0:4] == "UID:" or str(ws.cell(row=i, column=7).value)[0:4] == "uid:"):
            ws.cell(row=i, column=7).value = int(str(ws.cell(row=i, column=7).value)[4:])
        else:
            ws.cell(row=i, column=7).value = int(str(ws.cell(row=i, column=7).value))
    wb.save('修正数据.xlsx')
    wb.close()