import csv
import os
import openpyxl

if(__name__=="__main__"):
    wb = openpyxl.load_workbook('../修正数据.xlsx')
    ws = wb['Sheet1']

    header = []
    final_res = []
    for i in range(7, 39):
        header.append(str(ws.cell(row=1, column=i).value).strip('\n'))
    with open("../minimax_results_big_type/11254045.csv", 'r') as f:
        res = list(csv.reader(f))
        for i in range(1, len(res)):
            header.append("大类" + res[i][1])
    with open("../minimax_results/11254045.csv", 'r') as f:
        res = list(csv.reader(f))
        for i in range(1, len(res)):
            header.append("小类" + str(int(float(res[i][2]))))
    final_res.append(header)
    for i in range(2, ws.max_row + 1):
        if (not os.access("../minimax_results/" + str(ws.cell(row=i, column=7).value) + ".csv", os.F_OK)):
            continue
        res = []
        for j in range(7, 39):
            res.append(str(ws.cell(row=i, column=j).value))
        with open("../minimax_results_big_type/" + str(ws.cell(row=i, column=7).value) + ".csv", 'r') as f:
            res_ = list(csv.reader(f))
            for j in range(1, len(res_)):
                res.append(res_[j][2])
        with open("../minimax_results/" + str(ws.cell(row=i, column=7).value) + ".csv", 'r') as f:
            res_ = list(csv.reader(f))
            for j in range(1, len(res_)):
                res.append(res_[j][3])
        final_res.append(res)
    with open("../x_y_concatenation.csv", "w", newline='') as f:
        writer = csv.writer(f)
        writer.writerows(final_res)