import asyncio
import csv
import math
import os
import random
import aiohttp
import numpy
import numpy as np
import requests
import json
import openpyxl

from bs4 import BeautifulSoup
from tqdm import tqdm
from fake_useragent import UserAgent

tid_list = []
columns = ["UID", "大类", "小类", "投稿视频", "专栏分区", "收藏视频", "订阅视频", "投币视频", "点赞视频", "关注视频", "关注视频log2", "关注视频ln"]
proxies_set = []
max_error_num = 100
can_continue = False
finished = 0
temp_res = None
ua = UserAgent()

class TidList:
    def __init__(self, tid, name, type_tid, type):
        self.tid = tid
        self.name = name
        self.type_tid = type_tid
        self.type = type

    def __str__(self):
        return f"{self.tid} {self.name} {self.type_tid} {self.type}"

    def __eq__(self, other): return int(self.tid) == int(other.tid)

    def __le__(self, other): return int(self.tid) < int(other.tid)

    def __gt__(self, other): return int(self.tid) > int(other.tid)

def formalize(js):
    return json.dumps(js, indent=4, ensure_ascii=False)

def find_column(tid):
    l = 0
    r = len(tid_list) - 1
    while(l != r):
        mid = int((l + r) / 2)
        if(int(tid) == int(tid_list[mid].tid)):
            return mid
        elif(int(tid) < int(tid_list[mid].tid)):
            r = mid - 1
        else:
            l = mid + 1
    return l

async def get_random_proxy():

    proxypool_url = 'http://47.254.248.134:5555/random'
    session = aiohttp.ClientSession()
    res = await session.get(proxypool_url)
    res = await res.text()
    await session.close()
    return res.strip()

async def get_and_validate(url, params, cookies):

    global max_error_num, proxies_set

    failure = True
    true_res = None
    session = None
    error_num = 0
    base_prob = 20

    while(failure):
        headers = {
            'User-Agent': ua.random
        }
        rand = random.randint(0, 100)
        if(len(proxies_set) == 0 or rand > base_prob):
            proxy = 'http://' + await get_random_proxy()
        else:
            proxy = random.choice(proxies_set)
        # print(proxy)
        # print(proxies_set)

        try:
            session = aiohttp.ClientSession()
            res = await session.get(url, headers=headers, proxy=proxy, params=params, cookies=random.choice(cookies), timeout=2)
            res = await res.text()
            res = json.loads(res)
            await session.close()
            print(res)
            if (res["code"] == -412):
                # print("请求被拦截！")
                error_num += 1
                base_prob += 20
                if(proxy in proxies_set):
                    proxies_set.remove(proxy)
                if(error_num >= max_error_num):
                    print("被拦截次数过多，爬取失败")
                    exit(-1)
            else:
                true_res = res
                failure = False
                if (proxy not in proxies_set):
                    proxies_set.append(proxy)
        except:
            # print("代理链接超时！")
            await session.close()
            base_prob += 20
            if (proxy in proxies_set):
               proxies_set.remove(proxy)

    return true_res

async def get_masterpiece(page_num):
    url = "http://api.bilibili.com/x/space/arc/search"
    params = {
        "mid": str(mid),
        "pn": str(page_num),
        "ps": "50",
    }
    res = await get_and_validate(url, params, cookies)
    if(len(res["data"]["list"]["vlist"]) == 0):
        global can_continue
        can_continue = True
    else:
        for item in res["data"]["list"]["vlist"]:
            # print(formalize(item))
            result[find_column(item["typeid"])][3] += 1
    global finished
    finished += 1
    print(f"Task {finished} finished!")

async def get_collect(page_num, item):
    url = "http://api.bilibili.com/x/v3/fav/resource/list"
    params = {
        "media_id": item["id"],
        "pn": str(page_num),
        "ps": "20",
    }
    res = await get_and_validate(url, params, cookies)
    if (res["data"]["medias"] == None):
        global can_continue
        can_continue = True
    else:
        for item_j in res["data"]["medias"]:
            url = "http://api.bilibili.com/x/web-interface/view"
            params = {
                "bvid": item_j["bvid"]
            }
            res = await get_and_validate(url, params, cookies)
            try:
                # print(res["data"]["title"])
                # print(res["data"]["tid"])
                # print(res["data"]["desc"])
                # print(tid_list[find_column(res["data"]["tid"])])
                result[find_column(res["data"]["tid"])][5] += 1
            except:
                pass
    global finished
    finished += 1
    print(f"Task {finished} finished!")

async def get_collect_():
    url = "http://api.bilibili.com/x/v3/fav/folder/created/list-all"
    params = {
        "up_mid": str(mid),
    }
    res = await get_and_validate(url, params, cookies)
    global temp_res
    temp_res = res

async def get_collect__():
    url = "http://api.bilibili.com/x/v3/fav/folder/collected/list"
    params = {
        "up_mid": str(mid),
        "pn": "1",
        "ps": "50",
    }
    res = await get_and_validate(url, params, cookies)
    global temp_res
    temp_res = res

async def get_subscribe(page_num):
    url = "https://api.bilibili.com/x/space/bangumi/follow/list"
    params = {
        "vmid": str(mid),
        "type": "1",
        "pn": str(page_num),  # list为空则代表遍历完了
        "ps": "30"
    }
    res = await get_and_validate(url, params, cookies)
    global can_continue
    try:
        if (len(res["data"]["list"]) == 0):
            can_continue = True
        else:
            for item in res["data"]["list"]:
                result[find_column(item["season_type"])][6] += 1
    except:
        print("用户设置了隐私，无法访问")
        can_continue = True
        for i in range(len(tid_list)):
            result[i][6] = -1
    global finished
    finished += 1
    print(f"Task {finished} finished!")

async def get_coin_and_like():
    url = "http://api.bilibili.com/x/space/coin/video"
    params = {
        "vmid": str(mid),
    }
    res = await get_and_validate(url, params, cookies)
    global can_continue
    try:
        for item in res["data"]:
            result[find_column(item["tid"])][7] += 1
    except:
        print("用户设置了隐私，无法访问")
        can_continue = True
        for i in range(len(tid_list)):
            result[i][7] = -1
    url = "http://api.bilibili.com/x/space/like/video"
    params = {
        "vmid": "11254045",
    }
    res = await get_and_validate(url, params, cookies)
    try:
        for item in res["data"]["list"]:
            result[find_column(item["tid"])][8] += 1
    except:
        print("用户设置了隐私，无法访问")
        can_continue = True
        for i in range(len(tid_list)):
            result[i][8] = -1

async def get_follow(page_num, item):
    url = "http://api.bilibili.com/x/space/arc/search"
    params = {
        "mid": item["mid"],
        "pn": str(page_num),
        "ps": "50",
    }
    res = await get_and_validate(url, params, cookies)
    if (len(res["data"]["list"]["vlist"]) == 0):
        global can_continue
        can_continue = True
    else:
        for item in res["data"]["list"]["vlist"]:
            # print(item["title"])
            # print(item["typeid"])
            # print(item["description"])
            result_[find_column(item["typeid"])] += 1
    global finished
    finished += 1
    print(f"Task {finished} finished!")

async def get_follow_(i):
    url = "http://api.bilibili.com/x/relation/followings"
    params = {
        "vmid": str(mid),
        "pn": str(i),  # 限制只访问前5页（抽样，可以允许）
        "ps": "50",
    }
    res = await get_and_validate(url, params, cookies)
    global temp_res
    temp_res = res

if(__name__=="__main__"):

    cookies = [
        {"SESSDATA": "*"},
        {"SESSDATA": "*"},
        {"SESSDATA": "*"},
        {"SESSDATA": "*"},
        {"SESSDATA": "*"},
    ]

    # region 获取tid
    # 分区代码 详见https://github.com/SocialSisterYi/bilibili-API-collect/blob/master/video/video_zone.md
    url = "https://github.com/SocialSisterYi/bilibili-API-collect/blob/master/video/video_zone.md"
    html = BeautifulSoup(requests.get(url).text, 'lxml')
    div_text = html.find('div', id='readme')
    tr_text = div_text.findAll('tr')
    have_type = False
    have_type_tid = 0
    have_type_name = ""
    for item in tr_text:
        td_text = item.findAll('td')
        # print(td_text)
        if(len(td_text) == 0):
            have_type = False
            have_type_tid = 0
            have_type_name = ""
        else:
            delete_text = td_text[0].findAll('del')
            if(len(delete_text) != 0 and td_text[2].text != "124" and td_text[2].text != "95"):
                continue
            name = td_text[0].text
            tid = td_text[2].text
            if (tid == "124"):
                name = "社科·法律·心理"
            if (tid == "95"):
                name = "数码"
            if (tid == "76"):
                name = "美食制作"
            if(not have_type):
                have_type = True
                have_type_tid = tid
                have_type_name = name.split('(')[0]
            tid_list.append(TidList(tid, name, have_type_tid, have_type_name))
    tid_list = sorted(tid_list)
    for item in tid_list:
        print(item)
    # endregion

    # region 获取用户mid
    result = numpy.zeros([len(tid_list), len(columns)])
    wb = openpyxl.load_workbook('../修正数据.xlsx')
    ws = wb['Sheet1']

    mid_list = []
    for i in range(2, ws.max_row + 1):
        if(os.access(f"results/{ws.cell(row=i, column=7).value}.csv", os.F_OK)):
            continue
        else:
            mid_list.append(ws.cell(row=i, column=7).value)
    # endregion

    for o in range(len(mid_list)):
        mid = str(mid_list[o])
        print(f"爬取用户：{mid}")

        for i in range(len(tid_list)):
            for j in range(3, 12):
                result[i][j] = 0

        # region 获取用户投稿视频分区
        print("正在获取用户投稿视频分区")
        can_continue = False
        page_num = 1
        step = 20
        finished = 0
        while (page_num < 1000):
            tasks = [asyncio.ensure_future(get_masterpiece(page_num + _)) for _ in range(step)]
            loop = asyncio.get_event_loop()
            loop.run_until_complete(asyncio.wait(tasks))
            if(can_continue):
                break
            page_num += step
        print("获取完毕")
        # endregion
        '''
        # region 获取用户投稿动态分区（专栏api没找到）
        url = "http://api.bilibili.com/x/space/album/index"
        data = {
            "mid": "238547115",
            "ps": "50"
        }
        res = json.loads(requests.get(url, params=data).text)
        # print(formalize(res))
        print(formalize(res))
        # endregion
        '''
        # region 获取用户收藏夹视频分区
        print("正在获取用户收藏夹视频分区")
        access1 = False
        access2 = False
        tasks = [asyncio.ensure_future(get_collect_())]
        loop = asyncio.get_event_loop()
        loop.run_until_complete(asyncio.wait(tasks))
        try:
            for item in temp_res["data"]["list"]:
                can_continue = False
                page_num = 1
                step = 20
                finished = 0
                while (page_num < 1000):
                    tasks = [asyncio.ensure_future(get_collect(page_num + _, item)) for _ in range(step)]
                    loop = asyncio.get_event_loop()
                    loop.run_until_complete(asyncio.wait(tasks))
                    if (can_continue):
                        break
                    page_num += step
        except:
            print("用户设置了隐私，无法访问")
            access1 = True

        tasks = [asyncio.ensure_future(get_collect__())]
        loop = asyncio.get_event_loop()
        loop.run_until_complete(asyncio.wait(tasks))
        try:
            for item in temp_res["data"]["list"]:
                can_continue = False
                page_num = 1
                step = 20
                finished = 0
                while (page_num < 1000):
                    tasks = [asyncio.ensure_future(get_collect(page_num + _, item)) for _ in range(step)]
                    loop = asyncio.get_event_loop()
                    loop.run_until_complete(asyncio.wait(tasks))
                    if (can_continue):
                        break
                    page_num += step
        except:
            print("用户设置了隐私，无法访问")
            access2 = True
        if(access1 and access2):
            for i in range(len(tid_list)):
                result[i][5] = -1
        print("获取完毕")
        # endregion
        # region 获取用户订阅视频分区
        print("正在获取用户订阅视频分区")
        can_continue = False
        page_num = 1
        step = 20
        finished = 0
        while (page_num < 1000):
            tasks = [asyncio.ensure_future(get_subscribe(page_num + _)) for _ in range(step)]
            loop = asyncio.get_event_loop()
            loop.run_until_complete(asyncio.wait(tasks))
            if(can_continue):
                break
            page_num += step
        print("获取完毕")
        # endregion
        # region 获取用户投币点赞视频分区
        print("正在获取用户投币点赞视频分区")
        tasks = [asyncio.ensure_future(get_coin_and_like())]
        loop = asyncio.get_event_loop()
        loop.run_until_complete(asyncio.wait(tasks))
        print("获取完毕")
        # endregion
        # region 获取用户关注列表
        print("正在获取关注列表")
        up_step = 0
        for i in range(5):
            tasks = [asyncio.ensure_future(get_follow_(i + 1))]
            loop = asyncio.get_event_loop()
            loop.run_until_complete(asyncio.wait(tasks))
            try:
                for item in temp_res["data"]["list"]:
                    up_step += 1
                    print(print(up_step))
                    print(item["uname"])
                    can_continue = False
                    page_num = 1
                    step = 20
                    finished = 0
                    while (page_num < 1000):
                        result_ = np.zeros([len(tid_list)])
                        tasks = [asyncio.ensure_future(get_follow(page_num + _, item)) for _ in range(step)]
                        loop = asyncio.get_event_loop()
                        loop.run_until_complete(asyncio.wait(tasks))
                        for i in range(len(tid_list)):
                        #     print(result_[i])
                            result[i][9] += result_[i]
                            result[i][10] += math.log2(result_[i] + 1)
                            result[i][11] += math.log(result_[i] + 1)
                        #     print(result[i][9], result[i][10], result[i][11])
                        if (can_continue):
                            break
                        page_num += step
            except:
                print("用户设置了隐私，无法访问")
                can_continue = True
                for i in range(len(tid_list)):
                    result[i][9] = -1
                    result[i][10] = -1
                    result[i][11] = -1
        print("获取完毕")
        # endregion

        with open(f'results/{mid}.csv', 'w', newline='') as f:
            writer = csv.writer(f)
            writer.writerow(columns)
            for i in range(len(tid_list)):
                result[i][0] = int(mid)
                result[i][1] = tid_list[i].type_tid
                result[i][2] = tid_list[i].tid
                writer.writerow(result[i])
        print(f"完成对用户{mid}的扫描")